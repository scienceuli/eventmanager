import os
import openpyxl
import zipfile


from openpyxl.styles import NamedStyle

from datetime import datetime
from django.shortcuts import render, redirect
from django.utils.timezone import now
from django.http import HttpResponse
from django.db.models import DecimalField, FloatField, IntegerField
from django.contrib import admin, messages
from decimal import Decimal

from invoices.models import Invoice
from django.core.files.storage import default_storage

from utilities.pdf import render_to_pdf_directly

from .forms import SetDateForm


def export_to_excel_short(modeladmin, request, queryset):
    response = export_to_excel(modeladmin, request, queryset, short=True)
    return response


def export_to_excel(modeladmin, request, queryset, short=False):
    opts = modeladmin.model._meta
    model_name = opts.model_name
    # decimal_style = NamedStyle(name="decimal_style", number_format="0,00")
    decimal_style = NamedStyle(name="decimal_style", number_format="###0.00")
    wb = openpyxl.Workbook()
    if "decimal_style" not in wb.named_styles:
        wb.add_named_style(decimal_style)
    ws = wb.active
    ws.title = "Rechnungen Export" if model_name == "standardinvoice" else "Stornorechnungen Export"
    
    fields_no_export = [
        "uuid",
    ]

    if short:
        if model_name == "standardinvoice":
            fields = [
                "invoice_number",
                "get_full_name_and_events",
                "amount",
                "invoice_date",
            ]
        elif model_name == "stornoinvoice":
            fields = [
                "invoice_number",
                "get_full_name_and_events",
                "get_storno_amount",
                "invoice_date",
            ]
    else:
        fields = [
            field
            for field in opts.get_fields()
            if not field.many_to_many
            and not field.one_to_many
            and field.name not in fields_no_export
        ]

    # Define the header row
    if short:
        headers = [
            "Belegfeld",
            "Buchungstext",
            "Umsatz",
            "Datum",
            "Konto",
            "Gegenkonto",
            "KOST1",
            "S/H Kennzeichen",
        ]
    else:
        headers = [field.verbose_name for field in fields]
        headers.append("Betrag")
    ws.append(headers)

    # Append data rows
    for obj in queryset:
        data_row = []
        numeric_columns = []
        if short:
            for col_idx, field in enumerate(fields, start=1):
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    # If it's a method, call it
                    value = value() if callable(value) else value
                    # print(
                    #     f"Field: {field}, Value: {value}, Type: {type(value)}"
                    # )  # Debugging line
                    if isinstance(value, datetime):
                        value = value.strftime("%d.%m.%Y")
                    if isinstance(
                        value, (int, float, DecimalField, FloatField, Decimal)
                    ):
                        value = (
                            float(value) if value is not None else 0.00
                        )  # Convert to float
                        numeric_columns.append(col_idx)
                    data_row.append(value)
            data_row.extend(["10000", "8000", "4", "S"])
        else:
            for field in fields:
                value = getattr(obj, field.name)
                if isinstance(value, datetime):
                    value = value.strftime("%d.%m.%Y")
                data_row.append(value)
                data_row.append(obj.get_total_cost())
        ws.append(data_row)
        last_row = ws.max_row
        for col_idx in numeric_columns:
            ws.cell(row=last_row, column=col_idx).style = decimal_style

        obj.invoice_export = datetime.now()
        obj.save()

    filename = f"{opts.verbose_name}_{datetime.today().strftime('%Y-%m-%d')}"
    if short:
        filename = filename + "_kurz"

    # Prepare the response
    content_disposition = f"attachment; filename={filename}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = content_disposition

    wb.save(response)

    return response


def export_pdfs_as_zip(modeladmin, request, queryset):
    # Create in-memory ZIP file
    from io import BytesIO
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for invoice in queryset:
            if invoice.pdf:
                file_path = invoice.pdf.name
                file_name = os.path.basename(file_path)

                # Read file from storage
                with default_storage.open(file_path, 'rb') as f:
                    file_data = f.read()
                    zip_file.writestr(file_name, file_data)
                    
                invoice.pdf_export = datetime.now()
                invoice.save()
    
    zip_buffer.seek(0)

    model_name = modeladmin.model._meta.model_name

    storno_prefix = 'Storno-' if model_name == 'stornoinvoice' else ''

    filename = f"{storno_prefix}invoices_{datetime.today():%Y-%m-%d}.zip"
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def set_date_action(modeladmin, request, queryset):
    """
    An action that allows the user to set a date for a given queryset of objects.

    The user is presented with a form that allows them to select a date. If the form is valid,
    the selected date is applied to all objects in the queryset.
    """
    if 'apply' in request.POST:
        form = SetDateForm(request.POST)
        if form.is_valid():
            selected_date = form.cleaned_data['date']
            updated = queryset.update(invoice_receipt=selected_date)
            modeladmin.message_user(request, f"{updated} Rechnungen wurden aktualisiert.")
            return redirect(request.get_full_path())
    else:
        form = SetDateForm(initial={'date': now().date()})

    return render(request, 'admin/set_date_action.html', {
        'items': queryset,
        'form': form,
    })

set_date_action.short_description = "Rechnungseingang setzen"


def set_paid_action(modeladmin, request, queryset):
    """
    An action that allows the user to set a given queryset of objects as paid.

    The user is presented with a form that allows them to select a date. If the form is valid,
    the selected date is applied to all objects in the queryset.
    """
    now_date = now().date()
    updated =queryset.update(invoice_receipt=now_date)
    
    modeladmin.message_user(request, f"{updated} Rechnungen wurden aktualisiert.")
    return redirect(request.get_full_path())

set_paid_action.short_description = "Bezahlstatus setzen"