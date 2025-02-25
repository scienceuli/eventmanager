import csv
import zipfile
import openpyxl
from openpyxl.styles import NamedStyle

from datetime import datetime, date
from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.db.models import DecimalField, FloatField, IntegerField
from decimal import Decimal

from shop.models import OrderItem

from utilities.pdf import render_to_pdf_directly

from payment.utils import (
    generate_zip,
    update_order,
    check_order_date_in_future,
)


def export_to_csv(modeladmin, request, queryset):
    opts = modeladmin.model._meta
    content_disposition = f"attachment; filename={opts.verbose_name}_{datetime.today().strftime('%Y-%m-%d')}.csv"
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = content_disposition
    writer = csv.writer(response)
    fields = [
        field
        for field in opts.get_fields()
        if not field.many_to_many and not field.one_to_many
    ]
    # Write a first row with header information
    writer.writerow([field.verbose_name for field in fields])
    # Write data rows
    for obj in queryset:
        data_row = []
        for field in fields:
            value = getattr(obj, field.name)
            if isinstance(value, datetime):
                value = value.strftime("%d.%m.%Y")
            data_row.append(value)
        writer.writerow(data_row)
    return response


def export_to_excel_short(modeladmin, request, queryset):
    response = export_to_excel(modeladmin, request, queryset, short=True)
    return response


def export_to_excel(modeladmin, request, queryset, short=False):
    # decimal_style = NamedStyle(name="decimal_style", number_format="0,00")
    decimal_style = NamedStyle(name="decimal_style", number_format="###0.00")
    wb = openpyxl.Workbook()
    if "decimal_style" not in wb.named_styles:
        wb.add_named_style(decimal_style)
    ws = wb.active
    ws.title = "Rechnungen Export"
    opts = modeladmin.model._meta
    fields_no_export = ["uuid", "country"]

    if short:
        fields = [
            "get_order_number",
            "get_full_name_and_events",
            "get_total_cost",
            "payment_date",
            "payment_receipt",
        ]
    else:
        fields = [
            field
            for field in opts.get_fields()
            if not field.many_to_many
            and not field.one_to_many
            and field.name not in fields_no_export
        ]

    # Define the header row
    if short:
        headers = [
            "Belegfeld",
            "Buchungstext",
            "Umsatz",
            "Datum",
            "Konto",
            "Gegenkonto",
            "S/H Kennzeichen",
        ]
    else:
        headers = [field.verbose_name for field in fields]
        headers.append("Betrag")
    ws.append(headers)

    # Append data rows
    for obj in queryset:
        data_row = []
        numeric_columns = []
        if short:
            for col_idx, field in enumerate(fields, start=1):
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    # If it's a method, call it
                    value = value() if callable(value) else value
                    # print(
                    #     f"Field: {field}, Value: {value}, Type: {type(value)}"
                    # )  # Debugging line
                    if isinstance(value, datetime):
                        value = value.strftime("%d.%m.%Y")
                    if isinstance(
                        value, (int, float, DecimalField, FloatField, Decimal)
                    ):
                        value = (
                            float(value) if value is not None else 0.00
                        )  # Convert to float
                        numeric_columns.append(col_idx)
                    data_row.append(value)
            data_row.extend(["10000", "8000", "S"])
        else:
            for field in fields:
                value = getattr(obj, field.name)
                if isinstance(value, datetime):
                    value = value.strftime("%d.%m.%Y")
                data_row.append(value)
                data_row.append(obj.get_total_cost())
        ws.append(data_row)
        last_row = ws.max_row
        for col_idx in numeric_columns:
            ws.cell(row=last_row, column=col_idx).style = decimal_style

    filename = f"{opts.verbose_name}_{datetime.today().strftime('%Y-%m-%d')}"
    if short:
        filename = filename + "_kurz"

    # Prepare the response
    content_disposition = f"attachment; filename={filename}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = content_disposition

    wb.save(response)

    return response


def download_invoices_as_zipfile(modeladmin, request, queryset):
    zipfile_name = f"rechnungen_{datetime.today().strftime('%Y-%m-%d')}.zip"

    template_path = "shop/pdf_invoice.html"
    files = []

    for q in queryset.filter(download_marker=False):
        if check_order_date_in_future(q):
            update_order(q)
        context = {"order": q}
        context["order_items"] = OrderItem.objects.filter(order=q, status="r")
        context["contains_action_price"] = any(
            [
                item.is_action_price
                for item in OrderItem.objects.filter(order=q, status="r")
            ]
        )
        pdf = render_to_pdf_directly(template_path, context)
        filename = "rechnung_%s" % (q.get_order_number)
        files.append((filename + ".pdf", pdf))
        q.download_marker = True
        q.save()

    full_zip_in_memory = generate_zip(files)

    response = HttpResponse(
        full_zip_in_memory, content_type="application/force-download"
    )
    response["Content-Disposition"] = 'attachment; filename="{}"'.format(zipfile_name)

    return response


def reset_download_markers(modeladmin, request, queryset):
    for obj in queryset:
        obj.download_marker = False
        obj.save()
